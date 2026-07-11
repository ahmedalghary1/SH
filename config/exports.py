import csv
from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from config.pdf_utils import arabic_paragraph, register_arabic_font, shape_arabic


def resolve_value(obj, accessor):
    if callable(accessor):
        return accessor(obj)
    value = obj
    for part in accessor.split('.'):
        value = getattr(value, part, '')
        if callable(value):
            value = value()
        if value is None:
            return ''
    return value


def build_export_rows(queryset, columns):
    headers = [label for label, _accessor in columns]
    rows = []
    for obj in queryset:
        rows.append([resolve_value(obj, accessor) for _label, accessor in columns])
    return headers, rows


def export_csv_response(*, filename, headers, rows):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def export_xlsx_response(*, filename, title, headers, rows, metadata=()):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'كشف الحساب'
    sheet.sheet_view.rightToLeft = True
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet['A1'].font = Font(bold=True, size=16)
    for label, value in metadata:
        sheet.append([label, value])
    sheet.append([])
    header_row = sheet.max_row + 1
    sheet.append(headers)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='071D35')
    for row in rows:
        sheet.append(list(row))
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='right')
    for column_index, column in enumerate(sheet.columns, start=1):
        # The title row is merged across all columns. Cells inside that merged
        # range are MergedCell instances and do not expose ``column_letter``.
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(12, max(len(str(cell.value or '')) for cell in column) + 2),
            45,
        )
    for row in range(header_row + 1, sheet.max_row + 1):
        date_cell = sheet.cell(row, 1)
        if isinstance(date_cell.value, datetime):
            date_cell.number_format = 'dd/mm/yyyy hh:mm'
        elif isinstance(date_cell.value, date):
            date_cell.number_format = 'dd/mm/yyyy'
        for col in range(5, min(8, len(headers) + 1)):
            sheet.cell(row, col).number_format = '#,##0.00'
    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def _p(value, style):
    if isinstance(value, datetime):
        value = timezone.localtime(value).strftime('%d/%m/%Y %H:%M') if timezone.is_aware(value) else value.strftime('%d/%m/%Y %H:%M')
    elif isinstance(value, date):
        value = value.strftime('%d/%m/%Y')
    return arabic_paragraph(value, style)


def export_pdf_response(*, filename, title, headers, rows):
    font_name = register_arabic_font()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ExportTitle',
        parent=styles['Title'],
        fontName=font_name,
        fontSize=15,
        leading=21,
        alignment=TA_RIGHT,
    )
    meta_style = ParagraphStyle(
        'ExportMeta',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=9,
        leading=13,
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#475569'),
    )
    cell_style = ParagraphStyle(
        'ExportCell',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=8,
        leading=11,
        alignment=TA_RIGHT,
    )
    header_style = ParagraphStyle(
        'ExportHeader',
        parent=cell_style,
        textColor=colors.white,
        fontSize=8.5,
    )

    table_rows = [[_p(header, header_style) for header in reversed(headers)]]
    for row in rows:
        table_rows.append([_p(value, cell_style) for value in reversed(row)])
    if not rows:
        table_rows.append([_p('لا توجد بيانات', cell_style)] + [''] * (len(headers) - 1))

    story = [
        Paragraph(shape_arabic(title), title_style),
        Paragraph(shape_arabic(f'تاريخ التصدير: {timezone.localtime().strftime("%Y-%m-%d %H:%M")}'), meta_style),
        Paragraph(shape_arabic(f'عدد السجلات: {len(rows)}'), meta_style),
        Spacer(1, 8),
    ]
    table = Table(table_rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#071D35')),
        ('GRID', (0, 0), (-1, -1), .25, colors.HexColor('#DDE5EE')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    doc.build(story)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


class ExportListMixin:
    export_title = 'تصدير'
    export_filename = 'export'
    export_columns = ()

    def dispatch(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format in {'excel', 'pdf'}:
            return self.export_response(export_format)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['export_enabled'] = bool(self.export_columns)
        return context

    def export_response(self, export_format):
        queryset = self.get_queryset()
        headers, rows = build_export_rows(queryset, self.export_columns)
        if export_format == 'pdf':
            return export_pdf_response(
                filename=self.export_filename,
                title=self.export_title,
                headers=headers,
                rows=rows,
            )
        return export_csv_response(
            filename=self.export_filename,
            headers=headers,
            rows=rows,
        )
