from io import BytesIO
from datetime import date, time
from decimal import Decimal

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from accounts.models import User
from customers.models import Customer
from finance.models import CashAccount, PaymentTransaction
from orders.models import Order


class CustomerFeatureTests(TestCase):
    def setUp(self):
        self.manager = User.objects.create_user(username='customer-manager', password='x', role=User.ROLE_MANAGER)
        self.customer = Customer.objects.create(name='عميل الاختبار', address='القاهرة المعادي', opening_balance=100, created_by=self.manager)
        self.client.force_login(self.manager)

    def test_search_includes_address(self):
        response = self.client.get(reverse('customers:list'), {'q': 'المعادي'})
        self.assertContains(response, self.customer.name)

    def test_statement_exports_real_xlsx(self):
        response = self.client.get(reverse('customers:statement_export', args=[self.customer.pk, 'xlsx']))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b'PK'))
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        header_row = next(row for row in sheet.iter_rows() if row[0].value == 'التاريخ')
        first_data_cell = sheet.cell(header_row[0].row + 1, 1)
        self.assertIsInstance(first_data_cell.value, date)
        self.assertEqual(first_data_cell.number_format, 'dd/mm/yyyy')
        self.assertIn('عليه', values)
        self.assertIn('له', values)
        self.assertNotIn('المدين', values)
        self.assertNotIn('الدائن', values)

    def test_manager_can_delete_customer(self):
        response = self.client.post(reverse('customers:delete', args=[self.customer.pk]))

        self.assertRedirects(response, reverse('customers:simple_list'))
        self.assertFalse(Customer.objects.filter(pk=self.customer.pk).exists())

    def test_sales_user_cannot_delete_customer(self):
        sales_user = User.objects.create_user(username='customer-sales', password='x', role=User.ROLE_SALES)
        self.client.force_login(sales_user)

        response = self.client.post(reverse('customers:delete', args=[self.customer.pk]))

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Customer.objects.filter(pk=self.customer.pk).exists())

    def test_delete_action_is_visible_to_manager(self):
        response = self.client.get(reverse('customers:simple_list'))

        self.assertContains(response, reverse('customers:delete', args=[self.customer.pk]))

    def test_customer_accounts_show_latest_sale_and_receipt_values(self):
        order = Order.objects.create(
            order_number='ORD-CUSTOMER-LATEST',
            order_type=Order.TYPE_B2C,
            customer=self.customer,
            status=Order.STATUS_COMPLETED,
            total=Decimal('325.50'),
            created_by=self.manager,
        )
        cash = CashAccount.objects.create(name='Customer latest receipt cash')
        PaymentTransaction.objects.create(
            transaction_type=PaymentTransaction.TYPE_CUSTOMER_PAYMENT,
            direction=PaymentTransaction.DIRECTION_IN,
            amount=Decimal('125.25'),
            cash_account=cash,
            related_customer=self.customer,
            transaction_date=date(2026, 8, 16),
            transaction_time=time(14, 35),
            created_by=self.manager,
        )

        response = self.client.get(reverse('customers:simple_list'))
        listed_customer = list(response.context['customers'])[0]

        self.assertEqual(listed_customer.last_sale_at, order.created_at)
        self.assertEqual(listed_customer.last_invoice_value, Decimal('325.50'))
        self.assertEqual(listed_customer.last_receipt_date, date(2026, 8, 16))
        self.assertEqual(listed_customer.last_receipt_time, time(14, 35))
        self.assertEqual(listed_customer.last_receipt_amount, Decimal('125.25'))
        self.assertContains(response, 'قيمة آخر فاتورة')
        self.assertContains(response, 'قيمة آخر قبض')
