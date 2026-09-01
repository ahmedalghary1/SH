from dataclasses import dataclass, field
from decimal import Decimal, DecimalException, InvalidOperation
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
from django.db import DataError, IntegrityError, transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config.search import normalize_arabic

from .models import Product, ProductVariant


MAX_IMPORT_ROWS = 5000
MAX_PRICE = Decimal('99999999.99')
INVISIBLE_TEXT_MARKS = '\ufeff\u200b\u200c\u200d\u200e\u200f\u202a\u202b\u202c\u202d\u202e'


class ProductImportFileError(ValueError):
    pass


@dataclass
class ProductImportResult:
    created_count: int = 0
    skipped_count: int = 0
    errors: list[str] = field(default_factory=list)


def _normalize_arabic_digits(value):
    return str(value).translate(str.maketrans('٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹', '01234567890123456789'))


def _clean_text(value):
    return str(value).translate(str.maketrans('', '', INVISIBLE_TEXT_MARKS)).strip()


def _sku_duplicate_key(value):
    return _normalize_arabic_digits(_clean_text(value)).casefold()


def _name_duplicate_key(value):
    return normalize_arabic(_clean_text(value))


def _normalized_header(value):
    text = _normalize_arabic_digits(_clean_text(value or '')).replace('\u0640', '')
    text = ' '.join(text.split())
    aliases = {
        'م': 'serial',
        'م.': 'serial',
        'الكود': 'sku',
        'كود': 'sku',
        'اسم الصنف': 'name',
        'اسم المنتج': 'name',
        'سعر القطعة الرئيسي': 'cost_price',
        'سعر القطعه الرئيسي': 'cost_price',
        'سعر القطعة الرئيسى': 'cost_price',
        'سعر القطعه الرئيسى': 'cost_price',
        'الجملة': 'wholesale_price',
        'سعر الجملة': 'wholesale_price',
        'القطاعي': 'retail_price',
        'القطاعى': 'retail_price',
        'سعر القطاعي': 'retail_price',
        'سعر القطاعى': 'retail_price',
        'السنه': 'season',
        'السنة': 'season',
    }
    return aliases.get(text)


def _cell_text(cell, *, max_length, normalize_digits=False):
    value = cell.value
    if value is None:
        return ''
    if isinstance(value, bool):
        text = str(value)
    elif isinstance(value, int):
        text = str(value)
        number_format = (cell.number_format or '').split(';', 1)[0]
        if number_format and set(number_format) == {'0'}:
            text = text.zfill(len(number_format))
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = _clean_text(value)
    text = _clean_text(text)
    if normalize_digits:
        text = _normalize_arabic_digits(text)
    return text[:max_length + 1]


def _parse_price(value):
    if value is None or value == '':
        raise ValueError('السعر مطلوب')
    if isinstance(value, bool):
        raise ValueError('السعر غير صحيح')
    if isinstance(value, (int, float, Decimal)):
        raw = str(value)
    else:
        raw = _normalize_arabic_digits(value).strip()
        raw = raw.replace('٬', ',').replace('٫', '.')
        if ',' in raw and '.' not in raw and raw.count(',') == 1 and len(raw.rsplit(',', 1)[1]) <= 2:
            raw = raw.replace(',', '.')
        else:
            raw = raw.replace(',', '')
        raw = raw.replace(' ', '')
    try:
        price = Decimal(raw)
    except (InvalidOperation, ValueError):
        raise ValueError('السعر غير صحيح')
    if not price.is_finite() or price < 0 or price > MAX_PRICE:
        raise ValueError('السعر يجب أن يكون بين 0 و 99,999,999.99')
    try:
        return price.quantize(Decimal('0.01'))
    except DecimalException:
        raise ValueError('السعر غير صحيح')


def _variant_sku_for(product_sku):
    base = f'{product_sku}-0-0'
    candidate = base
    counter = 2
    while ProductVariant.objects.filter(variant_sku=candidate).exists():
        candidate = f'{base}-{counter}'
        counter += 1
    return candidate


def import_products_workbook(uploaded_file):
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True, keep_links=False)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError, EOFError, ParseError) as exc:
        raise ProductImportFileError('تعذر قراءة الملف. تأكد أنه ملف Excel سليم بصيغة XLSX أو XLSM.') from exc

    try:
        sheet = workbook.active
        rows = sheet.iter_rows()
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ProductImportFileError('ملف Excel فارغ.') from exc

        columns = {}
        for index, cell in enumerate(header_row):
            key = _normalized_header(cell.value)
            if key and key not in columns:
                columns[key] = index

        required = {'serial', 'sku', 'name', 'cost_price', 'wholesale_price', 'retail_price', 'season'}
        missing = required - columns.keys()
        if missing:
            raise ProductImportFileError(
                'عناوين الصف الأول غير صحيحة. الترتيب المطلوب من اليمين: '
                'م - الكود - اسم الصنف - سعر القطعة الرئيسي - الجملة - القطاعي - السنه.'
            )

        parsed_rows = []
        file_skus = set()
        file_names = set()
        result = ProductImportResult()

        for excel_row_number, row in enumerate(rows, start=2):
            if excel_row_number > MAX_IMPORT_ROWS + 1:
                raise ProductImportFileError(f'الملف يحتوي على أكثر من {MAX_IMPORT_ROWS} صف بيانات.')
            if not any(cell.value not in (None, '') for cell in row):
                continue

            def cell_for(key):
                index = columns[key]
                return row[index] if index < len(row) else None

            sku_cell = cell_for('sku')
            name_cell = cell_for('name')
            sku = _cell_text(sku_cell, max_length=100, normalize_digits=True) if sku_cell else ''
            name = _cell_text(name_cell, max_length=200) if name_cell else ''
            season_cell = cell_for('season')
            season = _cell_text(season_cell, max_length=100, normalize_digits=True) if season_cell else ''

            row_errors = []
            if not sku:
                row_errors.append('الكود مطلوب')
            elif len(sku) > 100:
                row_errors.append('الكود أطول من 100 حرف')
            if not name:
                row_errors.append('اسم الصنف مطلوب')
            elif len(name) > 200:
                row_errors.append('اسم الصنف أطول من 200 حرف')
            if len(season) > 100:
                row_errors.append('السنة أطول من 100 حرف')

            sku_key = _sku_duplicate_key(sku) if sku else ''
            name_key = _name_duplicate_key(name) if name else ''
            if sku_key and sku_key in file_skus:
                row_errors.append('الكود مكرر داخل الملف')
            if name_key and name_key in file_names:
                row_errors.append('اسم الصنف مكرر داخل الملف')

            try:
                cost_price = _parse_price(cell_for('cost_price').value if cell_for('cost_price') else None)
            except ValueError as exc:
                cost_price = None
                row_errors.append(f'سعر القطعة الرئيسي: {exc}')
            try:
                wholesale_price = _parse_price(cell_for('wholesale_price').value if cell_for('wholesale_price') else None)
            except ValueError as exc:
                wholesale_price = None
                row_errors.append(f'سعر الجملة: {exc}')
            try:
                retail_price = _parse_price(cell_for('retail_price').value if cell_for('retail_price') else None)
            except ValueError as exc:
                retail_price = None
                row_errors.append(f'سعر القطاعي: {exc}')

            if row_errors:
                result.skipped_count += 1
                result.errors.append(f'الصف {excel_row_number}: ' + '، '.join(row_errors))
                continue

            file_skus.add(sku_key)
            file_names.add(name_key)
            parsed_rows.append({
                'row_number': excel_row_number,
                'sku': sku,
                'name': name,
                'season': season or None,
                'cost_price': cost_price,
                'wholesale_price': wholesale_price,
                'retail_price': retail_price,
            })

        if not parsed_rows and not result.errors:
            raise ProductImportFileError('لا توجد بيانات منتجات تحت صف العناوين.')

        existing_products = Product.objects.values_list('sku', 'name')
        existing_skus = {_sku_duplicate_key(sku) for sku, _ in existing_products}
        existing_names = {_name_duplicate_key(name) for _, name in existing_products}
        new_rows = []
        for values in parsed_rows:
            duplicate_reasons = []
            if _sku_duplicate_key(values['sku']) in existing_skus:
                duplicate_reasons.append('الكود موجود بالفعل في النظام')
            if _name_duplicate_key(values['name']) in existing_names:
                duplicate_reasons.append('اسم الصنف موجود بالفعل في النظام')
            if duplicate_reasons:
                result.skipped_count += 1
                result.errors.append(f"الصف {values['row_number']}: " + '، '.join(duplicate_reasons))
            else:
                new_rows.append(values)
        parsed_rows = new_rows

        for values in parsed_rows:
            try:
                # Each row has its own transaction. A bad row is rolled back
                # completely without cancelling products imported before it.
                with transaction.atomic():
                    product = Product.objects.create(
                        name=values['name'],
                        sku=values['sku'],
                        season=values['season'],
                        retail_price=values['retail_price'],
                        wholesale_price=values['wholesale_price'],
                    )
                    ProductVariant.objects.create(
                        product=product,
                        variant_sku=_variant_sku_for(product.sku),
                        cost_price=values['cost_price'],
                        sale_price=values['retail_price'],
                        retail_price=values['retail_price'],
                        wholesale_price=values['wholesale_price'],
                    )
            except (IntegrityError, DataError, ValidationError, ValueError):
                result.skipped_count += 1
                result.errors.append(
                    f"الصف {values['row_number']}: تعذر حفظ المنتج؛ راجع الكود والأسعار ثم حاول مجددًا"
                )
            else:
                result.created_count += 1

        return result
    finally:
        workbook.close()
